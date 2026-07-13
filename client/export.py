"""
export.py — XLSX export of selected samples.

One summary row per requested sample (kept counts per channel, coexpressing
pairs, rate) plus one detail row per kept cell across every channel that has
data. Unlike earlier, a sample doesn't need every channel processed to be
exportable — each channel is handled independently, and any missing/not-yet-
processed piece renders as `NO_DATA` ("###") in the summary rather than
silently dropping the whole sample. This matters for large corpora: a reviewer
partway through processing still wants to export what's done so far, not wait
for every last sample and channel to finish first.

Reuses `coexpression.compute_coexpression` for the pair counts (only possible
when both CCK and CHR have data), so the export and the on-screen Composite
view can never disagree.

XLSX only (via `openpyxl` directly, no `pandas` — matching this project's
stance on keeping dependencies light, e.g. `scikit-image` over `shapely` in
coexpression.py) — a single workbook holds both tables as sheets, which a
two-file CSV export couldn't do as cleanly, and simplifies the export dialog
to one format decision fewer.
"""

from __future__ import annotations

from pathlib import Path
from typing import Callable

import openpyxl

import coexpression
from manifest import Manifest

NO_DATA = "###"

SUMMARY_HEADERS = ["prefix", "snap_kept", "cck_kept", "chr_kept", "coexpressing_pairs",
                    "coexpression_rate_pct"]
CELL_HEADERS = ["prefix", "channel", "cell_id", "status", "area", "centroid_x", "centroid_y",
                "source", "edited", "coexpressing"]


def _kept(cells: list[dict]) -> list[dict]:
    return [c for c in cells if c["status"] == "kept"]


def build_export_rows(prefixes: list[str], samples: dict[str, dict[str, str]], manifest: Manifest,
                       channel_status: Callable[[str, str], str]) -> tuple[list[dict], list[dict]]:
    """Returns (summary_rows, cell_rows) for exactly the given sample prefixes.

    `channel_status(prefix, channel) -> "ready" | "processing" | "not_ready"` is
    passed in rather than re-derived here so this stays in sync with whatever
    review.py already uses to light up the sidebar dots and gate the Composite tab.
    A channel that isn't `"ready"` (not processed, or doesn't exist for this
    sample) contributes no cell rows and shows as `NO_DATA` in the summary — a
    channel that *is* processed but genuinely detected zero kept cells still
    shows a real `0`, since "we don't know yet" and "we counted zero" are
    different, useful-to-distinguish facts.
    """
    summary_rows: list[dict] = []
    cell_rows: list[dict] = []

    for prefix in prefixes:
        channels = samples.get(prefix, {})
        kept_by_channel: dict[str, list[dict] | None] = {}
        for ch in ("SNAP", "CCK", "CHR"):
            filename = channels.get(ch)
            if filename is not None and channel_status(prefix, ch) == "ready":
                kept_by_channel[ch] = _kept(manifest.load_cells(filename))
            else:
                kept_by_channel[ch] = None

        snap_kept, cck_kept, chr_kept = kept_by_channel["SNAP"], kept_by_channel["CCK"], kept_by_channel["CHR"]

        if cck_kept is not None and chr_kept is not None:
            result = coexpression.compute_coexpression(cck_kept, chr_kept)
            n_pairs = len(result.pairs)
        else:
            result = None
            n_pairs = None

        if n_pairs is not None and snap_kept is not None and len(snap_kept) > 0:
            rate = round(n_pairs / len(snap_kept) * 100, 2)
        else:
            # Unknown pairs, unknown population, or a zero population (a rate
            # against zero cells is undefined, not zero) — all "###" alike.
            rate = None

        summary_rows.append({
            "prefix": prefix,
            "snap_kept": len(snap_kept) if snap_kept is not None else NO_DATA,
            "cck_kept": len(cck_kept) if cck_kept is not None else NO_DATA,
            "chr_kept": len(chr_kept) if chr_kept is not None else NO_DATA,
            "coexpressing_pairs": n_pairs if n_pairs is not None else NO_DATA,
            "coexpression_rate_pct": rate if rate is not None else NO_DATA,
        })

        for channel, cells, coexpr_ids in (
            ("SNAP", snap_kept, None),
            ("CCK", cck_kept, result.cck_ids if result else None),
            ("CHR", chr_kept, result.chr_ids if result else None),
        ):
            if cells is None:
                continue  # no data for this channel — nothing to list, ### already covers it above
            for c in cells:
                if channel == "SNAP":
                    coexpressing = ""  # not applicable — SNAP isn't part of the CCK/CHR overlap itself
                elif coexpr_ids is not None:
                    coexpressing = c["id"] in coexpr_ids
                else:
                    coexpressing = NO_DATA  # this channel has data but its CCK/CHR partner doesn't
                cell_rows.append({
                    "prefix": prefix,
                    "channel": channel,
                    "cell_id": c["id"],
                    "status": c["status"],
                    "area": c["area"],
                    "centroid_x": c["centroid"][0],
                    "centroid_y": c["centroid"][1],
                    "source": c["source"],
                    "edited": c.get("edited", False),
                    "coexpressing": coexpressing,
                })

    return summary_rows, cell_rows


def export_xlsx(path: Path, summary_rows: list[dict] | None, cell_rows: list[dict] | None) -> Path:
    """`summary_rows`/`cell_rows` of `None` omits that sheet entirely — callers
    are expected to guarantee at least one is not `None`."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # default blank sheet — only add what was actually requested

    if summary_rows is not None:
        ws = wb.create_sheet("Summary")
        ws.append(SUMMARY_HEADERS)
        for row in summary_rows:
            ws.append([row[h] for h in SUMMARY_HEADERS])

    if cell_rows is not None:
        ws = wb.create_sheet("Cells")
        ws.append(CELL_HEADERS)
        for row in cell_rows:
            ws.append([row[h] for h in CELL_HEADERS])

    wb.save(path)
    return path
